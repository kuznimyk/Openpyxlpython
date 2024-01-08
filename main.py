"""
Session: 1D01
Group Members: Mykyta Kuznietsov, Mark Hanson,
Due Date: Friday, Novermber 24, 2023
Assignment 3: Python introduction using openpyxl
Summary:takes data from the sells, puts the useful data into another sheet and creates a chart
Resources Used :
https://openpyxl.readthedocs.io/en/stable/"""



'''BEFORE RUNNING THE PROGRAM CREATE 4 EMPTY WORKSHEET IN THE EXCEL FILE'''



from openpyxl import *
from openpyxl.chart import BarChart,Reference

'''Mykyta Kuznietsov's part'''
'''function that creates chart where there is fixed about of data at x-axis like months and years'''
def fixedchartgenerator(datasheet):
    maxrow = datasheet.max_row
    monthchart = BarChart()
    monthchart.type = 'col'
    monthchart.style = 10
    monthchart.title = 'Collisions in each month '
    monthchart.x_axis_title = 'Month'
    monthchart.y_axis.title = 'Collisions'

    data = Reference(datasheet, min_col=2, max_col=2, min_row=1, max_row=maxrow)
    titles = Reference(datasheet, min_col=1, max_col=1, min_row=1, max_row=maxrow)
    monthchart.add_data(data, titles_from_data=False)
    monthchart.set_categories(titles)
    monthchart.width = 30
    monthchart.height = 15
    datasheet.add_chart(monthchart, "C1")

'''function that creates chart for varying amount of x-axis'''
def varyingchartgenerator(datasheet, collisioncount, maxrow):
    animalchart = BarChart()
    animalchart.type = 'col'
    animalchart.style = 10
    animalchart.title = 'Collision chart'
    animalchart.y_axis.title = 'Collisions'
    animalchart.x_axis.title = 'Animal'
    # animalchart.gapWidth = 2
    if collisioncount > 15:
        data = Reference(datasheet, min_col=4, max_col=4, min_row=1, max_row=maxrow)
        titles = Reference(datasheet, min_col=3, max_col=3, min_row=1, max_row=maxrow)
    else:
        data = Reference(datasheet, min_col=2, max_col=2, min_row=1, max_row=maxrow)
        titles = Reference(datasheet, min_col=1, max_col=1, min_row=1, max_row=maxrow)
    animalchart.add_data(data, titles_from_data=False)
    animalchart.set_categories(titles)
    animalchart.shape = 4
    animalchart.width = 30
    animalchart.height = 15
    datasheet.add_chart(animalchart, "E1")

'''Mark Hanson's part'''
def animalsdata(wb, datasheet):
    """part a"""
    #creating a chart for the animals
    animaldatasheet = wb.create_sheet('ChartForAnimals')
    #creating a dictionary animals are the keys and colisions are values
    collisions = {}
    #it is not a fixed so if there is more than 15, just use data from the top 10%
    collisionCounter = 0
    #skipping first row of data
    skip = True
    max = 1
    #looping thought the rows of data
    for row in datasheet:
        if skip == True:
            skip = False
            continue
        #get the name of the bird in each row
        if row[31].value:
            fullname = row[31].value
            #if the name has more than 2 words take the second one which will be the general name
            name = fullname.split()

            #keep the track of the total amount of different species
            if len(name) == 2:
                if name[1] in collisions:
                #if the animal is in the dictionary, than increase the value, if not - add animal to the dictionary with the value 1
                    collisions[name[1]] += 1
                    if collisions[name[1]] > max:
                        max = collisions[name[1]]
                else:
                    collisions[name[1]] = 1
                    collisionCounter += 1
            else:
                #checking for any unknowns
                if name[0] == 'UNKNOWN':
                    continue
                else:
                    if fullname in collisions:
                        collisions[fullname] += 1
                        if collisions[fullname] > max:
                            max = collisions[fullname]
                    else:
                        collisions[fullname] = 1
                        collisionCounter += 1
    # sorting the data to put the animals
    collisions = dict(sorted(collisions.items()))
    curcell = 1

    # if number of animals is less than 16 just use that data
    if collisionCounter < 16:
        for collision, count in collisions.items():
            animaldatasheet.append([collision, count])
    else:
        # if the amount of animals hit is more than 15 take out the top 90%, and use that cell to form the sheet
        for collision, count in collisions.items():
            if max * 0.1 < count:
                print(str(collision) + " " + str(count))
                animaldatasheet['C' + str(curcell)] = collision
                animaldatasheet['D' + str(curcell)] = count
                curcell += 1
            else:
                animaldatasheet.append([collision, count])

    # draw chart for the sorted data
    varyingchartgenerator(animaldatasheet, collisionCounter, curcell)



'''Mykyta Kuznietsov's part'''
'''funciton to sort out the data into new sheet'''
def yearsdata(wb,datasheet):
    '''part b'''
    #dictionary where years is a key and the amount of crashes that year is the value
    years = {}
    #skip the first row of datasheet as it contanins general information, but not the data itself
    skip = True

    #create a sheet for the years
    yeardatasheet = wb.create_sheet('ChartForYears')

    #looping through the rows in the sheet
    for row in datasheet:
        if skip == True:
            skip = False
            continue

        #taking the value of the year in  a row
        year = row[1].value
        #if the year exists in dictionary, then increase amount by one, if not then just add the year to the dictionary
        if year in years:
            years[year] += 1
        else:
            years[year] = 1

    #append the data into the years sheet from the dictionary
    for year, value in years.items():
        yeardatasheet.append([year, value])

    #and draw a chart for the sorted data
    fixedchartgenerator(yeardatasheet)


'''Mark Hanson's part'''
def monthsdata(wb,datasheet):
    '''part c'''
    monthsdatasheet = wb.create_sheet('ChartForMonths')

    #skipping first row in the sheet
    skip = True

    # creating a dictionary
    #keys are months and values are the crashes
    months = {}

    #looping through the rows in sheet
    for row in datasheet:
        if skip == True:
            skip = False
            continue
    #taking the value for months in a row
        month = row[2].value
    #if the month is already in the dict then add one to the value, if not then add the month into the dict
        if month in months:
            months[month] += 1
        else:
            months[month] = 1

    #sorting the dictionary
    months = dict(sorted(months.items()))

    #adding the data into the months sheet from the dict
    for month, value in months.items():
        monthsdatasheet.append([month, value])
    #creating the chart for the sorted data
    fixedchartgenerator(monthsdatasheet)



'''Mykyta Kuznietsov's part'''
'''Function that creates data and chart for the airlines'''
def airlinesdata(wb,datasheet):
    '''part d'''

    #create a new sheet for the airline collisions
    airlinesdatasheet = wb.create_sheet('ChartForAirlines')

    #skip the first row of the datasheet
    skip = True

    #dirctionary to count the amount of collisions for each airline
    airlines = {}

    #because amount of airlines is not fixed, if there are more than 15 airlines, a chart for 10% of max has to be created
    airlinecounter = 0
    max = 1

    #loop throught the rows
    for row in datasheet:
        if skip == True:
            skip = False
            continue

        #take the airline name cell
        airline = row[5].value

        #counting the amount of collisions for each airline and keeping the track of the total amount of airlines
        if airline in airlines:
            airlines[airline] += 1
            if airlines[airline] > max:
                max = airlines[airline]
        else:
            if airline == 'UNKNOWN':
                continue
            else:
                airlines[airline] = 1
                airlinecounter += 1

    #sort the dictionary by the name of the airlines
    airlines = dict(sorted(airlines.items()))
    curcell = 1

    #if amount of airlines if less than 16 then just make a chart out of that data
    if airlinecounter < 16:
        for airline, count in airlines.items():
            airlinesdatasheet.append([airline, count])
    else:
    #if the amount of airlines is more than 15 than take out top 90%, put that data in the separate cells and take data for the chart from those cells
        for airline, count in airlines.items():
            if max * 0.1 < count:
                airlinesdatasheet['C' + str(curcell)] = airline
                airlinesdatasheet['D' + str(curcell)] = count
                curcell += 1
            else:

                airlinesdatasheet.append([airline, count])


    #draw the chart for the sorted data
    varyingchartgenerator(airlinesdatasheet, airlinecounter,curcell)























if __name__ == '__main__':
    #load the workbook
    wb= load_workbook('mediumAircraftData.xlsx')

    #remove 4 last sheets from the workbook
    datasheet = wb.worksheets[0]
    removesheet4 = wb.worksheets[4]
    wb.remove(removesheet4)
    removesheet3 = wb.worksheets[3]
    wb.remove(removesheet3)
    removesheet2 = wb.worksheets[2]
    wb.remove(removesheet2)
    removesheet1 = wb.worksheets[1]
    wb.remove(removesheet1)

    #create sheets of data and charts for animals, years, months and airlines
    animalsdata(wb, datasheet)
    yearsdata(wb,datasheet)
    monthsdata(wb,datasheet)
    airlinesdata(wb,datasheet)

    #save the workbook file
    wb.save('mediumAircraftData.xlsx')

# See PyCharm help at https://www.jetbrains.com/help/pycharm/



'''sources
how to sort the dictionary: https://www.freecodecamp.org/news/sort-dictionary-by-value-in-python/
plotting the bar charts in openpyxl: https://www.pylenin.com/blogs/bar-charts-openpyxl/
run through data in excel sheet https://www.geeksforgeeks.org/reading-excel-file-using-python/
'''
